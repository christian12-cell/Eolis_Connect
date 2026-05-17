import io
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from ..database import get_db
from ..models import User, AIUsage, CreditRequest, InfrastructureCost, FinancialAuditLog, FinancialProjection
from ..deps import get_current_user, require_roles
from ..credit_service import FREE_CREDITS_ON_SIGNUP

def _client_ip(request: Request) -> str | None:
    fwd = request.headers.get("x-forwarded-for")
    return fwd.split(",")[0].strip() if fwd else str(request.client.host) if request.client else None

router = APIRouter(prefix="/finance", tags=["finance"])

FINANCE_ROLES = ("FINANCE_AGENT", "SYSTEM_ADMIN")


class InfraCostIn(BaseModel):
    category: str
    label: str
    amount_fcfa: float
    amount_usd: float


class ProjectionIn(BaseModel):
    period: str                           # YYYY-MM
    target_revenue: float
    target_clients: int = 0
    target_net_profit: Optional[float] = None
    target_margin_pct: Optional[float] = None
    notes: Optional[str] = None
    period: str
    invoice_url: Optional[str] = None


def _apply_date_filter(q, model_col, from_date, to_date):
    if from_date:
        try: q = q.filter(model_col >= datetime.strptime(from_date, "%Y-%m-%d"))
        except ValueError: pass
    if to_date:
        try:
            d = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.filter(model_col <= d)
        except ValueError: pass
    return q


@router.get("/dashboard")
def finance_dashboard(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date:   Optional[str] = Query(None, alias="to"),
    current_user: User = Depends(require_roles(*FINANCE_ROLES)),
    db: Session = Depends(get_db),
):
    # Revenue — approved credit requests
    req_q = db.query(CreditRequest).filter(CreditRequest.status == "approved")
    req_q = _apply_date_filter(req_q, CreditRequest.validated_at, from_date, to_date)
    approved = req_q.all()
    total_revenue = sum(r.amount_validated or 0 for r in approved)

    # AI costs
    ai_q = db.query(AIUsage)
    ai_q = _apply_date_filter(ai_q, AIUsage.created_at, from_date, to_date)
    usages = ai_q.all()
    total_ai_fcfa  = sum(u.cost_fcfa for u in usages)
    total_credits  = sum(getattr(u, "credits_cost", 0) or 0 for u in usages)

    # Infrastructure costs
    infra_q = db.query(InfrastructureCost)
    if from_date:
        try:
            ym = from_date[:7]
            infra_q = infra_q.filter(InfrastructureCost.period >= ym)
        except Exception: pass
    if to_date:
        try:
            ym = to_date[:7]
            infra_q = infra_q.filter(InfrastructureCost.period <= ym)
        except Exception: pass
    infra_costs = infra_q.all()
    total_infra_fcfa = sum(c.amount_fcfa for c in infra_costs)
    total_infra_usd  = sum(c.amount_usd  for c in infra_costs)

    # Pending credit requests
    pending = db.query(CreditRequest).filter(CreditRequest.status == "pending").all()
    pending_amount = sum(r.amount_declared for r in pending)

    # Nouveaux clients (info seulement — leur coût réel est déjà dans totalAiCostFcfa)
    new_clients_q = db.query(User).filter(User.role == "CLIENT")
    new_clients_q = _apply_date_filter(new_clients_q, User.created_at, from_date, to_date)
    new_clients_count = new_clients_q.count()

    # P&L
    total_costs  = total_ai_fcfa + total_infra_fcfa
    gross_profit = total_revenue - total_ai_fcfa
    net_profit   = total_revenue - total_costs
    usage_profit = total_credits - total_ai_fcfa
    margin_pct   = round((net_profit / total_revenue * 100), 1) if total_revenue > 0 else None

    return {
        "totalRevenue":        round(total_revenue,       2),
        "totalAiCostFcfa":     round(total_ai_fcfa,       4),
        "totalAiCostUsd":      round(total_ai_fcfa / 600, 6),
        "totalCreditsConsumed": round(total_credits,      2),
        "usageProfit":         round(usage_profit,        4),
        "totalInfraFcfa":      round(total_infra_fcfa,    2),
        "totalInfraUsd":       round(total_infra_usd,     4),
        "newClientsCount":     new_clients_count,
        "totalCosts":          round(total_costs,         2),
        "grossProfit":         round(gross_profit,        2),
        "netProfit":           round(net_profit,          2),
        "marginPct":           margin_pct,
        "approvedCount":       len(approved),
        "pendingCount":        len(pending),
        "pendingAmount":       round(pending_amount,      2),
        "infraBreakdown":    [
            {
                "id":          c.id,
                "category":    c.category,
                "label":       c.label,
                "amountFcfa":  c.amount_fcfa,
                "amountUsd":   c.amount_usd,
                "period":      c.period,
                "invoiceUrl":  c.invoice_url,
                "addedBy":     f"{c.added_by_user.first_name} {c.added_by_user.last_name}" if c.added_by_user else None,
                "createdAt":   c.created_at.isoformat() + "Z",
            } for c in sorted(infra_costs, key=lambda x: x.period, reverse=True)
        ],
    }


@router.get("/infra-costs")
def list_infra_costs(
    current_user: User = Depends(require_roles(*FINANCE_ROLES)),
    db: Session = Depends(get_db),
):
    costs = db.query(InfrastructureCost).order_by(InfrastructureCost.period.desc(), InfrastructureCost.created_at.desc()).all()
    return [
        {
            "id":         c.id,
            "category":   c.category,
            "label":      c.label,
            "amountFcfa": c.amount_fcfa,
            "amountUsd":  c.amount_usd,
            "period":     c.period,
            "invoiceUrl": c.invoice_url,
            "addedBy":    f"{c.added_by_user.first_name} {c.added_by_user.last_name}" if c.added_by_user else None,
            "createdAt":  c.created_at.isoformat() + "Z",
        } for c in costs
    ]


@router.post("/infra-costs", status_code=201)
def add_infra_cost(
    request: Request,
    body: InfraCostIn,
    current_user: User = Depends(require_roles("FINANCE_AGENT")),
    db: Session = Depends(get_db),
):
    cost = InfrastructureCost(
        category=body.category,
        label=body.label,
        amount_fcfa=body.amount_fcfa,
        amount_usd=body.amount_usd,
        period=body.period,
        invoice_url=body.invoice_url,
        added_by=current_user.id,
    )
    db.add(cost)
    db.flush()
    db.add(FinancialAuditLog(
        user_id=current_user.id, action="INFRA_COST_ADD", entity_id=cost.id,
        amount_fcfa=body.amount_fcfa, details=f"{body.category} — {body.label} ({body.period})",
        ip_address=_client_ip(request),
    ))
    db.commit()
    db.refresh(cost)
    return {"id": cost.id, "label": cost.label, "amountFcfa": cost.amount_fcfa}


@router.delete("/infra-costs/{cost_id}", status_code=204)
def delete_infra_cost(
    request: Request,
    cost_id: str,
    current_user: User = Depends(require_roles("FINANCE_AGENT")),
    db: Session = Depends(get_db),
):
    cost = db.query(InfrastructureCost).filter(InfrastructureCost.id == cost_id).first()
    if not cost:
        raise HTTPException(404)
    db.add(FinancialAuditLog(
        user_id=current_user.id, action="INFRA_COST_DELETE", entity_id=cost_id,
        amount_fcfa=cost.amount_fcfa, details=f"{cost.category} — {cost.label}",
        ip_address=_client_ip(request),
    ))
    db.delete(cost)
    db.commit()


@router.get("/audit-log")
def get_audit_log(
    search:    Optional[str] = Query(None),
    period:    Optional[str] = Query(None),
    from_date: Optional[str] = Query(None, alias="from"),
    to_date:   Optional[str] = Query(None, alias="to"),
    page:      int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    current_user: User = Depends(require_roles("SYSTEM_ADMIN")),
    db: Session = Depends(get_db),
):
    """Financial audit trail — SYSTEM_ADMIN only, immutable."""
    import math
    from datetime import timedelta, timezone
    from sqlalchemy import or_

    q = db.query(FinancialAuditLog).order_by(FinancialAuditLog.created_at.desc())

    now = datetime.now(timezone.utc)
    if period == "day":
        q = q.filter(FinancialAuditLog.created_at >= now - timedelta(days=1))
    elif period == "week":
        q = q.filter(FinancialAuditLog.created_at >= now - timedelta(days=7))
    elif period == "month":
        q = q.filter(FinancialAuditLog.created_at >= now - timedelta(days=30))
    elif period == "year":
        q = q.filter(FinancialAuditLog.created_at >= now - timedelta(days=365))

    if from_date:
        try:
            q = q.filter(FinancialAuditLog.created_at >= datetime.fromisoformat(from_date))
        except ValueError:
            pass
    if to_date:
        try:
            q = q.filter(FinancialAuditLog.created_at <= datetime.fromisoformat(to_date))
        except ValueError:
            pass

    if search:
        s = f"%{search}%"
        q = q.outerjoin(User, FinancialAuditLog.user_id == User.id).filter(
            or_(
                FinancialAuditLog.action.ilike(s),
                FinancialAuditLog.details.ilike(s),
                FinancialAuditLog.ip_address.ilike(s),
                FinancialAuditLog.entity_id.ilike(s),
                User.username.ilike(s),
                User.first_name.ilike(s),
                User.last_name.ilike(s),
            )
        )

    total = q.count()
    logs = q.offset((page - 1) * page_size).limit(page_size).all()

    def _client_info(entity_id: str | None, action: str):
        if not entity_id or not action.startswith("CREDIT"):
            return None, None
        req = db.query(CreditRequest).filter(CreditRequest.id == entity_id).first()
        if not req or not req.client:
            return None, None
        return f"{req.client.first_name} {req.client.last_name}", req.client.username

    items = [
        {
            "id":             l.id,
            "action":         l.action,
            "entityId":       l.entity_id,
            "amountFcfa":     l.amount_fcfa,
            "details":        l.details,
            "ipAddress":      l.ip_address,
            "createdAt":      l.created_at.isoformat() + "Z",
            "doneBy":         f"{l.user.first_name} {l.user.last_name}" if l.user else l.user_id,
            "doneByUsername": l.user.username if l.user else None,
            "role":           l.user.role if l.user else None,
            **dict(zip(
                ("clientName", "clientUsername"),
                _client_info(l.entity_id, l.action)
            )),
        } for l in logs
    ]

    return {
        "total":    total,
        "page":     page,
        "pageSize": page_size,
        "pages":    math.ceil(total / page_size) if total > 0 else 1,
        "items":    items,
    }


@router.get("/pnl")
def pnl_report(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date:   Optional[str] = Query(None, alias="to"),
    urgency:   Optional[str] = Query(None),
    current_user: User = Depends(require_roles(*FINANCE_ROLES)),
    db: Session = Depends(get_db),
):
    """Monthly P&L breakdown."""
    from ..models import Ticket
    req_q = db.query(CreditRequest).filter(CreditRequest.status == "approved")
    req_q = _apply_date_filter(req_q, CreditRequest.validated_at, from_date, to_date)
    approved = req_q.all()

    ai_q = db.query(AIUsage)
    ai_q = _apply_date_filter(ai_q, AIUsage.created_at, from_date, to_date)
    usages = ai_q.all()

    # Apply urgency filter on AI usages via ticket
    urgency_list = [u.strip() for u in urgency.split(",")] if urgency else None
    if urgency_list:
        filtered = []
        for u in usages:
            ticket = None
            if u.ticket_id:
                ticket = db.query(Ticket).filter(Ticket.id == u.ticket_id).first()
            elif u.bl_document_id:
                ticket = db.query(Ticket).filter(Ticket.bl_document_id == u.bl_document_id).first()
            if ticket and ticket.urgency in urgency_list:
                filtered.append(u)
        usages = filtered

    infra = db.query(InfrastructureCost).all()

    def _blank():
        return {"revenue": 0, "aiCost": 0, "infraCost": 0, "credits": 0, "newClients": 0}

    # Group by month
    months: dict = {}

    for r in approved:
        if not r.validated_at: continue
        key = r.validated_at.strftime("%Y-%m")
        months.setdefault(key, _blank())
        months[key]["revenue"] += r.amount_validated or 0

    for u in usages:
        key = u.created_at.strftime("%Y-%m")
        months.setdefault(key, _blank())
        months[key]["aiCost"]  += u.cost_fcfa
        months[key]["credits"] += float(getattr(u, "credits_cost", 0) or 0)

    for c in infra:
        key = c.period
        months.setdefault(key, _blank())
        months[key]["infraCost"] += c.amount_fcfa

    # Nouveaux clients par mois (info seulement, pas un coût P&L)
    new_clients_q = db.query(User).filter(User.role == "CLIENT")
    new_clients_q = _apply_date_filter(new_clients_q, User.created_at, from_date, to_date)
    for client in new_clients_q.all():
        key = client.created_at.strftime("%Y-%m")
        months.setdefault(key, _blank())
        months[key]["newClients"] += 1

    rows = []
    for month, v in sorted(months.items()):
        net = v["revenue"] - v["aiCost"] - v["infraCost"]
        rows.append({
            "month":       month,
            "revenue":     round(v["revenue"],   2),
            "aiCost":      round(v["aiCost"],    4),
            "infraCost":   round(v["infraCost"], 2),
            "totalCost":   round(v["aiCost"] + v["infraCost"], 2),
            "grossProfit": round(v["revenue"] - v["aiCost"], 2),
            "netProfit":   round(net,             2),
            "credits":     round(v["credits"],    2),
            "newClients":  v["newClients"],
            "marginPct":   round(net / v["revenue"] * 100, 1) if v["revenue"] > 0 else None,
        })

    return rows


@router.get("/pnl/export-xlsx")
def export_pnl_xlsx(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date:   Optional[str] = Query(None, alias="to"),
    urgency:   Optional[str] = Query(None),
    current_user: User = Depends(require_roles(*FINANCE_ROLES)),
    db: Session = Depends(get_db),
):
    """Export P&L report as formatted Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    # ── Reuse pnl_report logic ─────────────────────────────────────────────────
    from ..models import Ticket
    req_q = db.query(CreditRequest).filter(CreditRequest.status == "approved")
    req_q = _apply_date_filter(req_q, CreditRequest.validated_at, from_date, to_date)
    approved = req_q.all()

    ai_q = db.query(AIUsage)
    ai_q = _apply_date_filter(ai_q, AIUsage.created_at, from_date, to_date)
    usages = ai_q.all()

    urgency_list = [u.strip() for u in urgency.split(",")] if urgency else None
    if urgency_list:
        filtered = []
        for u in usages:
            ticket = None
            if u.ticket_id:
                ticket = db.query(Ticket).filter(Ticket.id == u.ticket_id).first()
            elif u.bl_document_id:
                ticket = db.query(Ticket).filter(Ticket.bl_document_id == u.bl_document_id).first()
            if ticket and ticket.urgency in urgency_list:
                filtered.append(u)
        usages = filtered

    infra = db.query(InfrastructureCost).all()

    months: dict = {}
    for r in approved:
        if not r.validated_at: continue
        key = r.validated_at.strftime("%Y-%m")
        months.setdefault(key, {"revenue": 0, "aiCost": 0, "infraCost": 0, "credits": 0})
        months[key]["revenue"] += r.amount_validated or 0
    for u in usages:
        key = u.created_at.strftime("%Y-%m")
        months.setdefault(key, {"revenue": 0, "aiCost": 0, "infraCost": 0, "credits": 0})
        months[key]["aiCost"]  += u.cost_fcfa
        months[key]["credits"] += float(getattr(u, "credits_cost", 0) or 0)
    for c in infra:
        key = c.period
        months.setdefault(key, {"revenue": 0, "aiCost": 0, "infraCost": 0, "credits": 0})
        months[key]["infraCost"] += c.amount_fcfa

    rows = []
    for month, v in sorted(months.items()):
        net = v["revenue"] - v["aiCost"] - v["infraCost"]
        rows.append({
            "month": month,
            "revenue":    round(v["revenue"],   2),
            "aiCost":     round(v["aiCost"],    4),
            "infraCost":  round(v["infraCost"], 2),
            "totalCost":  round(v["aiCost"] + v["infraCost"], 2),
            "grossProfit": round(v["revenue"] - v["aiCost"], 2),
            "netProfit":  round(net, 2),
            "credits":    round(v["credits"], 2),
            "marginPct":  round(net / v["revenue"] * 100, 1) if v["revenue"] > 0 else None,
        })

    USD_RATE = 600.0
    EUR_RATE = 655.957

    def to_usd(f): return round(f / USD_RATE, 2)
    def to_eur(f): return round(f / EUR_RATE, 2)

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport P&L"
    ws.sheet_properties.tabColor = "1B3A5C"

    BLUE       = "1B3A5C"
    LIGHT_BLUE = "4A8FC4"
    WHITE      = "FFFFFF"
    GREEN_BG   = "ECFDF5"
    RED_BG     = "FEF2F2"
    ALT_BG     = "F0F4F8"
    TOTAL_BG   = "0F2A47"

    thin = Side(style="thin", color="D1DAE3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color=None, size=10, italic=False) -> Font:
        return Font(name="Calibri", size=size, bold=bold, italic=italic,
                    color=color or "1A202C")

    def align(h="left", v="center", wrap=False) -> Alignment:
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    HEADERS = [
        "Mois", "Revenus (FCFA)", "Revenus ($)", "Revenus (€)",
        "Coûts IA (FCFA)", "Charges infra (FCFA)", "Coûts totaux (FCFA)",
        "Bénéfice brut (FCFA)", "Bénéfice net (FCFA)", "Bénéf. net ($)", "Bénéf. net (€)", "Marge %",
    ]
    NUM_COLS = len(HEADERS)

    # Row 1 — Title
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    c = ws["A1"]
    period_label = f"{from_date} → {to_date}" if from_date and to_date else "Toute la période"
    c.value = f"EOLIS CONNECT — Rapport Financier P&L  ·  {period_label}"
    c.font  = font(bold=True, color=WHITE, size=14)
    c.fill  = fill(BLUE)
    c.alignment = align("center")
    ws.row_dimensions[1].height = 36

    # Row 2 — Subtitle
    ws.merge_cells(f"A2:{get_column_letter(NUM_COLS)}2")
    c = ws["A2"]
    c.value = f"Exporté le {datetime.utcnow().strftime('%d/%m/%Y à %H:%M')} UTC  ·  {len(rows)} mois"
    c.font  = font(italic=True, color=LIGHT_BLUE, size=9)
    c.fill  = fill("E8EEF4")
    c.alignment = align("center")
    ws.row_dimensions[2].height = 18

    # Row 3 — Column headers
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = font(bold=True, color=WHITE, size=10)
        c.fill      = fill(LIGHT_BLUE)
        c.alignment = align("center", wrap=True)
        c.border    = border
    ws.row_dimensions[3].height = 40
    ws.freeze_panes = "A4"

    # Data rows
    for i, row in enumerate(rows):
        r = 4 + i
        net = row["netProfit"]
        row_fill = fill(GREEN_BG if net >= 0 else RED_BG) if i % 2 == 0 else fill("FFFFFF" if net >= 0 else "FFF5F5")

        values = [
            row["month"],
            row["revenue"],    to_usd(row["revenue"]),    to_eur(row["revenue"]),
            row["aiCost"],     row["infraCost"],           row["totalCost"],
            row["grossProfit"],row["netProfit"],           to_usd(row["netProfit"]), to_eur(row["netProfit"]),
            f"{row['marginPct']}%" if row["marginPct"] is not None else "—",
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill   = row_fill
            c.border = border
            if col == 1:
                c.font      = font(bold=True)
                c.alignment = align("center")
            elif col == NUM_COLS:
                c.font      = font(bold=(row["marginPct"] is not None and row["marginPct"] >= 50),
                                   color="15803D" if (row["marginPct"] or 0) >= 50 else
                                         "B45309" if (row["marginPct"] or 0) >= 0 else "DC2626")
                c.alignment = align("right")
            else:
                c.number_format = '#,##0.00'
                c.alignment     = align("right")
                if col in (9,) and net < 0:
                    c.font = font(bold=True, color="DC2626")
                elif col in (9,):
                    c.font = font(bold=True, color="15803D")
                else:
                    c.font = font()
        ws.row_dimensions[r].height = 20

    # Totals row
    if rows:
        tr = 4 + len(rows)
        t_rev   = sum(r["revenue"]    for r in rows)
        t_ai    = sum(r["aiCost"]     for r in rows)
        t_infra = sum(r["infraCost"]  for r in rows)
        t_cost  = sum(r["totalCost"]  for r in rows)
        t_gross = sum(r["grossProfit"]for r in rows)
        t_net   = sum(r["netProfit"]  for r in rows)
        t_marg  = round(t_net / t_rev * 100, 1) if t_rev > 0 else None

        totals = [
            "TOTAL",
            t_rev, to_usd(t_rev), to_eur(t_rev),
            t_ai,  t_infra, t_cost,
            t_gross, t_net, to_usd(t_net), to_eur(t_net),
            f"{t_marg}%" if t_marg is not None else "—",
        ]
        for col, val in enumerate(totals, 1):
            c = ws.cell(row=tr, column=col, value=val)
            c.fill      = fill(TOTAL_BG)
            c.font      = font(bold=True, color=WHITE, size=10)
            c.border    = border
            c.alignment = align("right" if col > 1 else "center")
            if col not in (1, NUM_COLS) and isinstance(val, float):
                c.number_format = '#,##0.00'
        ws.row_dimensions[tr].height = 24

        # Footer note
        ws.merge_cells(f"A{tr+2}:{get_column_letter(NUM_COLS)}{tr+2}")
        c = ws.cell(row=tr+2, column=1, value="Généré automatiquement par Eolis Connect — Ne pas modifier")
        c.font      = font(italic=True, color="9CA3AF", size=8)
        c.alignment = align("center")

    # Column widths
    col_widths = [12, 18, 13, 13, 18, 20, 18, 20, 20, 13, 13, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Stream response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"eolis-pnl-{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Projections économiques ────────────────────────────────────────────────────

@router.get("/projections")
def list_projections(
    current_user: User = Depends(require_roles(*FINANCE_ROLES)),
    db: Session = Depends(get_db),
):
    rows = db.query(FinancialProjection).order_by(FinancialProjection.period).all()
    return [
        {
            "id":              r.id,
            "period":          r.period,
            "targetRevenue":    r.target_revenue,
            "targetClients":    r.target_clients,
            "targetNetProfit":  r.target_net_profit,
            "targetMarginPct":  r.target_margin_pct,
            "notes":           r.notes,
            "createdBy":       f"{r.creator.first_name} {r.creator.last_name}" if r.creator else None,
            "updatedAt":       r.updated_at.isoformat() + "Z",
        } for r in rows
    ]


@router.post("/projections", status_code=201)
def upsert_projection(
    body: ProjectionIn,
    current_user: User = Depends(require_roles("FINANCE_AGENT")),
    db: Session = Depends(get_db),
):
    import re
    if not re.match(r"^\d{4}-\d{2}$", body.period):
        raise HTTPException(400, "Format période invalide (attendu: YYYY-MM)")
    existing = db.query(FinancialProjection).filter(FinancialProjection.period == body.period).first()
    if existing:
        existing.target_revenue    = body.target_revenue
        existing.target_clients    = body.target_clients
        existing.target_net_profit = body.target_net_profit
        existing.target_margin_pct = body.target_margin_pct
        existing.notes             = body.notes
        existing.updated_at        = datetime.utcnow()
        db.commit()
        db.refresh(existing)
        return {"id": existing.id, "period": existing.period}
    proj = FinancialProjection(
        period=body.period, target_revenue=body.target_revenue,
        target_clients=body.target_clients, target_net_profit=body.target_net_profit,
        target_margin_pct=body.target_margin_pct,
        notes=body.notes, created_by=current_user.id,
    )
    db.add(proj)
    db.commit()
    db.refresh(proj)
    return {"id": proj.id, "period": proj.period}


@router.delete("/projections/{projection_id}", status_code=204)
def delete_projection(
    projection_id: str,
    current_user: User = Depends(require_roles("FINANCE_AGENT")),
    db: Session = Depends(get_db),
):
    proj = db.query(FinancialProjection).filter(FinancialProjection.id == projection_id).first()
    if not proj:
        raise HTTPException(404)
    db.delete(proj)
    db.commit()


@router.get("/projections/comparison")
def projection_comparison(
    year:         Optional[str] = Query(None),
    from_period:  Optional[str] = Query(None, alias="from"),   # YYYY-MM
    to_period:    Optional[str] = Query(None, alias="to"),     # YYYY-MM
    current_user: User = Depends(require_roles(*FINANCE_ROLES)),
    db: Session = Depends(get_db),
):
    """Compare objectifs vs réalité mois par mois. Supporte year=YYYY ou from/to=YYYY-MM."""
    import calendar

    # Projections filtrées
    proj_q = db.query(FinancialProjection).order_by(FinancialProjection.period)
    if from_period:
        proj_q = proj_q.filter(FinancialProjection.period >= from_period[:7])
    if to_period:
        proj_q = proj_q.filter(FinancialProjection.period <= to_period[:7])
    elif year and not from_period:
        proj_q = proj_q.filter(FinancialProjection.period.like(f"{year}-%"))
    projections = {p.period: p for p in proj_q.all()}

    # Dates réelles
    if from_period:
        from_date = f"{from_period[:7]}-01"
    elif year:
        from_date = f"{year}-01-01"
    else:
        from_date = None

    if to_period:
        y, m = map(int, to_period[:7].split("-"))
        last = calendar.monthrange(y, m)[1]
        to_date = f"{to_period[:7]}-{last:02d}"
    elif year and not from_period:
        to_date = f"{year}-12-31"
    else:
        to_date = None

    req_q = db.query(CreditRequest).filter(CreditRequest.status == "approved")
    req_q = _apply_date_filter(req_q, CreditRequest.validated_at, from_date, to_date)
    approved = req_q.all()

    ai_q = db.query(AIUsage)
    ai_q = _apply_date_filter(ai_q, AIUsage.created_at, from_date, to_date)
    usages = ai_q.all()

    infra = db.query(InfrastructureCost)
    if from_period:
        infra = infra.filter(InfrastructureCost.period >= from_period[:7])
    if to_period:
        infra = infra.filter(InfrastructureCost.period <= to_period[:7])
    elif year and not from_period:
        infra = infra.filter(InfrastructureCost.period.like(f"{year}-%"))
    infra = infra.all()

    new_clients_q = db.query(User).filter(User.role == "CLIENT")
    new_clients_q = _apply_date_filter(new_clients_q, User.created_at, from_date, to_date)
    new_clients = new_clients_q.all()

    actuals: dict = {}
    def _ab(): return {"revenue": 0.0, "aiCost": 0.0, "infraCost": 0.0, "newClients": 0}

    for r in approved:
        if not r.validated_at: continue
        key = r.validated_at.strftime("%Y-%m")
        actuals.setdefault(key, _ab())
        actuals[key]["revenue"] += r.amount_validated or 0

    for u in usages:
        key = u.created_at.strftime("%Y-%m")
        actuals.setdefault(key, _ab())
        actuals[key]["aiCost"] += u.cost_fcfa

    for c in infra:
        actuals.setdefault(c.period, _ab())
        actuals[c.period]["infraCost"] += c.amount_fcfa

    for cl in new_clients:
        key = cl.created_at.strftime("%Y-%m")
        actuals.setdefault(key, _ab())
        actuals[key]["newClients"] += 1

    # Merge all periods
    all_periods = sorted(set(list(projections.keys()) + list(actuals.keys())))

    rows = []
    for period in all_periods:
        proj = projections.get(period)
        act  = actuals.get(period, _ab())

        net = act["revenue"] - act["aiCost"] - act["infraCost"]
        actual_margin = round(net / act["revenue"] * 100, 1) if act["revenue"] > 0 else None

        target_rev = proj.target_revenue if proj else None
        rev_var    = round(act["revenue"] - target_rev, 2) if target_rev is not None else None
        rev_pct    = round((act["revenue"] / target_rev - 1) * 100, 1) if target_rev and target_rev > 0 else None

        if target_rev is None:
            status = "no_target"
        elif rev_pct is None:
            status = "no_data"
        elif rev_pct >= 5:
            status = "exceeded"
        elif rev_pct >= -5:
            status = "on_track"
        elif rev_pct >= -20:
            status = "behind"
        else:
            status = "far_behind"

        rows.append({
            "period":          period,
            "target": {
                "revenue":    target_rev,
                "clients":    proj.target_clients if proj else None,
                "netProfit":  proj.target_net_profit if proj else None,
                "marginPct":  proj.target_margin_pct if proj else None,
                "notes":      proj.notes if proj else None,
                "projectionId": proj.id if proj else None,
            },
            "actual": {
                "revenue":    round(act["revenue"], 2),
                "aiCost":     round(act["aiCost"], 4),
                "infraCost":  round(act["infraCost"], 2),
                "netProfit":  round(net, 2),
                "marginPct":  actual_margin,
                "newClients": act["newClients"],
            },
            "variance": {
                "revenue":       rev_var,
                "revenuePct":    rev_pct,
                "clients":       (act["newClients"] - proj.target_clients) if proj else None,
                "netProfit":     round(net - proj.target_net_profit, 2) if proj and proj.target_net_profit is not None else None,
                "netProfitPct":  round((net / proj.target_net_profit - 1) * 100, 1) if proj and proj.target_net_profit and proj.target_net_profit != 0 else None,
                "status":        status,
            },
        })

    return rows
